#!/usr/bin/env python3
"""
import_zip_full.py
All-in-one importer:
- Streams Excel -> CSV (memory-safe)
- Processes CSV in chunks, validates rows
- Batched, multi-threaded inserts
- Resume support, ETA, logging
- CLI options



# basic run
python import_zip_full.py --file "D:\Parasya\Moms Milk\mom-milk-api\src\data\zipcodes.xlsx"

# with options
python import_zip_full.py --file "..." --batch 2000 --workers 6 --country SG --resume
"""

import argparse
import csv
import json
import logging
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from math import ceil
from typing import List, Tuple

import mysql.connector
import pandas as pd
from mysql.connector import Error
from openpyxl import load_workbook

# ----------------------------
# CONFIG (change as needed)
# ----------------------------
DB_HOST = "localhost"
DB_USER = "palqardev"
DB_PASSWORD = "Parasya@2025"
DB_NAME = "moms_milk_db"
TABLE_NAME = "zip_codes"

DEFAULT_BATCH_SIZE = 1000
DEFAULT_WORKERS = 4
STATE_FILE = "import_state.json"
LOG_FILE = "import-log.txt"
# If your Excel already is CSV, set Excel file path to None and CSV_PATH to your file
DEFAULT_FILE_PATH = r"D:\Parasya\Moms Milk\mom-milk-api\src\data\zipcodes.xlsx"

# Mapping from Excel header -> DB field
COLUMN_MAP = {
    "country code": "country",
    "postal code": "zipcode",
    "place name": "placeName",
    "latitude": "latitude",
    "longitude": "longitude"
}

REQUIRED_FIELDS = ["country", "zipcode", "placeName", "latitude", "longitude"]

# ----------------------------
# Logging setup
# ----------------------------
logger = logging.getLogger("importer")
logger.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s %(levelname)s: %(message)s")

fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
fh.setLevel(logging.INFO)
fh.setFormatter(formatter)
logger.addHandler(fh)

ch = logging.StreamHandler(sys.stdout)
ch.setLevel(logging.INFO)
ch.setFormatter(formatter)
logger.addHandler(ch)


# ----------------------------
# Helper: MySQL connection maker (for threads)
# ----------------------------
def make_connection():
    return mysql.connector.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        autocommit=False,
        connection_timeout=30
    )


# ----------------------------
# Excel -> CSV streaming converter (memory safe)
# ----------------------------
def excel_to_csv_streaming(xlsx_path: str, csv_path: str) -> None:
    """Convert Excel to CSV streaming rows to avoid loading all into memory."""
    logger.info("Converting Excel -> CSV (streaming)...")
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    with open(csv_path, "w", newline="", encoding="utf-8") as fout:
        writer = csv.writer(fout)
        first = True
        for row in ws.iter_rows(values_only=True):
            # row is a tuple
            if first:
                writer.writerow([str(c).strip() if c is not None else "" for c in row])
                first = False
            else:
                writer.writerow([("" if v is None else v) for v in row])

    wb.close()
    logger.info("Excel -> CSV conversion finished: %s", csv_path)


# ----------------------------
# Count CSV lines (fast-ish)
# ----------------------------
def count_csv_rows(csv_path: str) -> int:
    with open(csv_path, "r", encoding="utf-8", errors="ignore") as f:
        count = sum(1 for _ in f) - 1  # subtract header
    return max(0, count)


# ----------------------------
# Prepare insert; used by worker threads
# ----------------------------
INSERT_QUERY = f"""
INSERT INTO {TABLE_NAME} (country, zipcode, placeName, latitude, longitude)
VALUES (%s, %s, %s, %s, %s)
ON DUPLICATE KEY UPDATE
    country = VALUES(country),
    placeName = VALUES(placeName),
    latitude = VALUES(latitude),
    longitude = VALUES(longitude);
"""


def insert_batch_worker(batch: List[Tuple], worker_id: int) -> Tuple[int, str]:
    """
    Insert a batch using its own DB connection.
    Returns (num_rows, error_message_or_empty)
    """
    conn = None
    try:
        conn = make_connection()
        cur = conn.cursor()
        cur.executemany(INSERT_QUERY, batch)
        conn.commit()
        cur.close()
        return len(batch), ""
    except Exception as e:
        logger.exception("Worker %s: insert error", worker_id)
        # try to rollback
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
        return 0, str(e)
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass


# ----------------------------
# Utility: Validate and transform a CSV row (dict)
# ----------------------------
def validate_and_map_row(row: dict) -> Tuple[bool, Tuple]:
    """
    Map CSV row to DB tuple and validate.
    Returns (is_valid, tuple_or_error_string)
    """
    try:
        # Normalize keys lower-case trimmed to match COLUMN_MAP earlier if needed
        mapped = {}
        for k_src, k_dst in COLUMN_MAP.items():
            # find actual key in provided row case-insensitively
            for actual in row.keys():
                if actual.strip().lower() == k_src.strip().lower():
                    mapped[k_dst] = row[actual]
                    break
            else:
                # if not found, keep None; we'll catch missing required fields below
                mapped[k_dst] = None

        # required checks
        if not mapped.get("zipcode") or str(mapped.get("zipcode")).strip() == "":
            return False, "missing zipcode"

        place = mapped.get("placeName") or ""
        if str(place).strip() == "":
            return False, "missing placeName"

        # latitude / longitude numeric
        lat = mapped.get("latitude")
        lon = mapped.get("longitude")
        try:
            lat_f = float(lat) if (lat is not None and str(lat).strip() != "") else None
            lon_f = float(lon) if (lon is not None and str(lon).strip() != "") else None
        except Exception:
            return False, "invalid lat/lon"

        # If lat/lon are missing, you may choose to skip; here we require them
        if lat_f is None or lon_f is None:
            return False, "missing lat/lon"

        country = mapped.get("country") or ""
        return True, (str(country).strip(), str(mapped["zipcode"]).strip(), str(place).strip(), lat_f, lon_f)

    except Exception as e:
        return False, f"exception {e}"


# ----------------------------
# Resume state helpers
# ----------------------------
def load_state() -> dict:
    if os.path.isfile(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_state(state: dict):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f)


# ----------------------------
# Main import routine
# ----------------------------
def run_import(csv_path: str, batch_size: int, workers: int, country_filter: str, resume: bool):
    # pre-checks
    if not os.path.isfile(csv_path):
        logger.error("CSV file not found: %s", csv_path)
        sys.exit(1)

    # MySQL connection quick-check
    try:
        test_conn = make_connection()
        test_conn.close()
    except Exception as e:
        logger.error("Cannot connect to MySQL: %s", e)
        sys.exit(1)

    total_rows = count_csv_rows(csv_path)
    logger.info("Total rows in CSV: %d", total_rows)
    if total_rows == 0:
        logger.info("No rows to import, exiting.")
        return

    # load state
    state = load_state() if resume else {}
    last_processed = state.get("last_processed", 0) if resume else 0
    logger.info("Resuming: %s, last_processed rows: %d", resume, last_processed)

    # compute starting place
    rows_to_skip = last_processed

    # prepare CSV reader in chunks using pandas
    # pandas read_csv chunksize yields DataFrame objects
    start_time = time.time()
    processed = last_processed
    batches_total_est = ceil((total_rows - processed) / batch_size)

    logger.info("Batch size: %d, workers: %d, estimated batches remaining: %d", batch_size, workers, batches_total_est)

    # We will create a ThreadPoolExecutor and submit batch insertion jobs
    executor = ThreadPoolExecutor(max_workers=workers)

    # iterate with pandas in chunks but we need to skip the initial rows if resuming
    chunk_iter = pd.read_csv(csv_path, chunksize=batch_size, iterator=True, encoding="utf-8", dtype=str, low_memory=True)

    # fast-skip by consuming chunks until we reach resume point
    if rows_to_skip > 0:
        to_skip_chunks = rows_to_skip // batch_size
        for _ in range(to_skip_chunks):
            try:
                next(chunk_iter)
            except StopIteration:
                break
        # if there are leftover rows to skip inside a chunk, we'll read one chunk and drop rows
        leftover = rows_to_skip % batch_size
        if leftover:
            try:
                df_first = next(chunk_iter)
                df_first = df_first.iloc[leftover:]
                # we'll process df_first as the first chunk by re-injecting it below
                first_chunk_df = df_first
            except StopIteration:
                first_chunk_df = None
        else:
            first_chunk_df = None
    else:
        first_chunk_df = None

    futures = []
    batch_index = processed // batch_size  # number of FULL batches already processed
    rows_done_including_previous = processed

    def submit_df(df_chunk):
        nonlocal batch_index, rows_done_including_previous, futures
        # transform df_chunk rows to validated tuples
        data_batch = []
        errors = 0
        for _, row in df_chunk.iterrows():
            is_valid, out = validate_and_map_row(row.to_dict())
            if not is_valid:
                logger.debug("Skipping row (invalid): %s", out)
                errors += 1
                continue
            # country filter
            if country_filter and out[0].strip().upper() != country_filter.strip().upper():
                continue
            data_batch.append(out)

        if not data_batch:
            # nothing to insert for this chunk
            batch_index += 1
            rows_done_including_previous += len(df_chunk)
            return

        # submit worker
        worker_id = batch_index % workers
        future = executor.submit(insert_batch_worker, data_batch, worker_id)
        futures.append((future, len(data_batch)))
        batch_index += 1
        rows_done_including_previous += len(df_chunk)

    # If we had a first_chunk_df from resume leftover, process it first
    if first_chunk_df is not None:
        submit_df(first_chunk_df)

    # iterate rest chunks
    for df_chunk in chunk_iter:
        submit_df(df_chunk)

        # if we have too many pending futures, wait for some to finish to avoid memory growth
        while len([f for f, _ in futures if not f.done()]) > workers * 2:
            # wait for first future to complete
            done_list = [f for f, _ in futures if f.done()]
            for done_f in done_list:
                # handle results
                for (fut, count) in list(futures):
                    if fut is done_f:
                        num_inserted, err = fut.result()
                        if err:
                            logger.error("Batch insert error: %s", err)
                        else:
                            logger.info("Inserted %d rows (batch).", num_inserted)
                        futures.remove((fut, count))
                        # update state
                        processed += count
                        state["last_processed"] = processed
                        save_state(state)
                        break

            # small sleep to avoid busy loop
            time.sleep(0.1)

        # progress logging (estimate ETA)
        elapsed = time.time() - start_time
        if processed > 0:
            rate = processed / elapsed
            remaining = max(0, total_rows - processed)
            eta_s = remaining / rate if rate > 0 else 0
        else:
            rate = 0
            eta_s = 0
        percent = round((processed / total_rows) * 100, 2) if total_rows else 0
        logger.info("Progress: %d/%d rows (%.2f%%), ETA: %s", processed, total_rows, percent, time.strftime("%H:%M:%S", time.gmtime(eta_s)))

    # after submitting all, wait for all futures to finish and update state
    for fut, count in futures:
        num_inserted, err = fut.result()
        if err:
            logger.error("Batch insert error: %s", err)
        else:
            logger.info("Inserted %d rows (batch final).", num_inserted)
        processed += count
        state["last_processed"] = processed
        save_state(state)

    executor.shutdown(wait=True)
    elapsed_total = time.time() - start_time
    logger.info("Import complete. Total processed rows: %d. Time elapsed: %s", processed, time.strftime("%H:%M:%S", time.gmtime(elapsed_total)))
    # clear state on success
    try:
        if os.path.isfile(STATE_FILE):
            os.remove(STATE_FILE)
            logger.info("Removed resume state file.")
    except Exception:
        pass


# ----------------------------
# CLI and orchestration
# ----------------------------
def main():
    parser = argparse.ArgumentParser(description="Import zipcodes excel/csv to MySQL with resume/ETA/logging.")
    parser.add_argument("--file", "-f", default=DEFAULT_FILE_PATH, help="Path to Excel (.xlsx) or CSV file")
    parser.add_argument("--batch", "-b", type=int, default=DEFAULT_BATCH_SIZE, help="Batch (chunk) size")
    parser.add_argument("--workers", "-w", type=int, default=DEFAULT_WORKERS, help="Parallel worker threads for DB inserts")
    parser.add_argument("--country", "-c", default="", help="Filter by country code (e.g., SG) - optional")
    parser.add_argument("--resume", "-r", action="store_true", help="Resume from last state if available")
    parser.add_argument("--convert-only", action="store_true", help="If Excel, convert to CSV only and exit")
    args = parser.parse_args()

    input_path = args.file
    # If it's an xlsx, convert it to CSV in same folder with same name .csv
    if input_path.lower().endswith(".xlsx") or input_path.lower().endswith(".xls"):
        csv_path = os.path.splitext(input_path)[0] + ".csv"
        if not os.path.isfile(csv_path):
            excel_to_csv_streaming(input_path, csv_path)
        else:
            logger.info("CSV already exists: %s", csv_path)
        if args.convert_only:
            logger.info("Conversion completed; exiting as --convert-only was passed.")
            return
    else:
        csv_path = input_path

    # quick header check: open the CSV and read headers, ensure required mapped fields exist
    with open(csv_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        headers = next(reader)
    headers_lc = [h.strip().lower() for h in headers]
    missing = []
    for src in COLUMN_MAP.keys():
        if src.strip().lower() not in headers_lc:
            missing.append(src)
    if missing:
        logger.warning("Some expected excel headers not found in CSV: %s", missing)
        logger.info("Available headers: %s", headers)
        # we continue because validate_and_map_row can still find keys case-insensitively,
        # but warn the user.

    run_import(csv_path, args.batch, args.workers, args.country, args.resume)


if __name__ == "__main__":
    main()
